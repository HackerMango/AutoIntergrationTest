import openpyxl as openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import xlrd
import sqlite3

database_FileName = r'RearWheel_APA.db'
src_ExcelFileName = "泊车.xlsx"
dir_ExcelFileName = "泊车" + "_1" + ".xlsx"

db_Connection = sqlite3.connect(database_FileName)
db_Cursor = db_Connection.cursor()
db_Cursor.execute('SELECT * from RxInfo_Table')
get_Data = db_Cursor.fetchall()
rxInfo_Table = [list(x) for x in get_Data]

db_Cursor.execute('SELECT * from Step_Table')
get_Data = db_Cursor.fetchall()
step_Table = [list(x) for x in get_Data]

ActSignalValue = []
Test_Result = []
Step_Name = []
temp = ""
current_StepName = ""
last_StepName = ""
x = rxInfo_Table.pop(0)

while len(rxInfo_Table):
    temp_rxInfo = rxInfo_Table.pop(0)
    if last_StepName == temp_rxInfo[3] or last_StepName == "":
        temp += temp_rxInfo[1] + "==" + str(temp_rxInfo[4]) + "\n"
    elif last_StepName != temp_rxInfo[3] or last_StepName != "":
        ActSignalValue.append(temp[: -1])
        Step_Name.append(last_StepName)
        temp = str()
    if len(rxInfo_Table) == 0:
        ActSignalValue.append(temp[: -1])
        Step_Name.append(last_StepName)
    last_StepName = temp_rxInfo[3]

for it_Step_Table in step_Table:
    Test_Result.append(it_Step_Table[3])
Excel_Book = xlrd.open_workbook(src_ExcelFileName)
Excel_Sheet = Excel_Book.sheets()
Test_Case_Data = []

nrows = Excel_Sheet[0].nrows
ncols = Excel_Sheet[0].ncols

for i in range(ncols):
    Test_Case_Data.append(Excel_Sheet[0].col_values(i)[:])

count_Test = 0
count_Step = 0
for row in range(nrows):
    if Step_Name[count_Step].find(Test_Case_Data[0][row]) != -1 and Test_Case_Data[1][row].find("用例") != -1:
        count_Test = row
        while (Step_Name[count_Step].find(Test_Case_Data[0][row]) != -1 and count_Step < len(Step_Name)):
            Test_Case_Data[10][count_Test + 1] = ActSignalValue[count_Step]
            Test_Case_Data[9][count_Test + 1] = Test_Result[count_Step]
            count_Test = count_Test + 1
            count_Step = count_Step + 1
            if Test_Case_Data[1][row].find("功能\nFunction") != -1:
                count_Test = count_Test + 1
                count_Step = count_Step + 1
            if count_Step >= len(Step_Name):
                break
        if count_Step >= len(Step_Name):
            break
        if count_Test > row:
            row = count_Test

work_Book = openpyxl.load_workbook(dir_ExcelFileName)
table = work_Book["Sheet1"]
table = work_Book.active

merged_col = []
merged_row = []
for it_bottom in table.merged_cells.ranges:
    if list(it_bottom.bounds)[1] not in merged_row:
        temp = list(it_bottom.bounds)[1]
        while temp <= list(it_bottom.bounds)[3]:
            merged_row.append(temp)
            temp = temp + 1

for i in range(1, len(Test_Case_Data[10]) + 1):
    if i not in merged_row:
        table.cell(i, 11).value = Test_Case_Data[10][i - 1]
        table.cell(i, 10).value = Test_Case_Data[9][i - 1]
        if isinstance(Test_Case_Data[9][i - 1], int):
            table.cell(i, 10).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if Test_Case_Data[9][i - 1] == -1:
                table.cell(i, 11).font = Font(color='c00000')
                table.cell(i, 10).value = "F"
            else:
                table.cell(i, 11).font = Font(color='000000')
                table.cell(i, 10).value = "P"

work_Book.save(dir_ExcelFileName)
